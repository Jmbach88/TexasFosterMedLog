"""
Tkinter GUI for Medication Tracker
Main application window with tabbed interface
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import json
from datetime import datetime
from pathlib import Path
import os
import sys
import subprocess
import platform
from PIL import Image, ImageTk

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.profiles import ProfileManager
from core.logs import LogManager
from core.export import ExportManager
from core.medication_cards import MedicationCardManager
from core.settings_manager import SettingsManager


class MedicationTrackerApp:
    """Main application window"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Medication Log Tracker")
        self.root.geometry("1000x700")
        self.root.minsize(900, 600)  # Set minimum window size

        # Initialize managers with error handling
        # Use ResourceManager (no data_dir parameter = uses ~/MedicationLogger/)
        try:
            self.profile_manager = ProfileManager()  # Uses ResourceManager
            self.data_dir = str(self.profile_manager.resource_manager.user_data_dir)
        except Exception as e:
            messagebox.showerror(
                "Initialization Error",
                f"Failed to initialize Profile Manager:\n{str(e)}\n\n"
                "The application may not function correctly."
            )
            # Use a fallback empty manager
            self.profile_manager = None
            self.data_dir = None

        try:
            self.log_manager = LogManager()  # Uses ResourceManager
        except Exception as e:
            messagebox.showerror(
                "Initialization Error",
                f"Failed to initialize Log Manager:\n{str(e)}\n\n"
                "Medication logging features will be unavailable."
            )
            self.log_manager = None

        try:
            self.medication_card_manager = MedicationCardManager()  # Uses ResourceManager
        except Exception as e:
            messagebox.showerror(
                "Initialization Error",
                f"Failed to initialize Medication Card Manager:\n{str(e)}\n\n"
                "Medication card features will be unavailable."
            )
            self.medication_card_manager = None

        try:
            self.export_manager = ExportManager()  # Uses ResourceManager
        except Exception as e:
            messagebox.showerror(
                "Initialization Error",
                f"Failed to initialize Export Manager:\n{str(e)}\n\n"
                "Export features will be unavailable.\n\n"
                "Please check that the template file is bundled with the application."
            )
            self.export_manager = None

        try:
            self.settings_manager = SettingsManager()
        except Exception as e:
            messagebox.showerror(
                "Initialization Error",
                f"Failed to initialize Settings Manager:\n{str(e)}\n\n"
                "Settings features will be unavailable."
            )
            self.settings_manager = None

        # Define standard fonts for consistent UI
        self.TITLE_FONT = ('Segoe UI', 12, 'bold')      # Main section titles
        self.HEADING_FONT = ('Segoe UI', 10, 'bold')    # Subsection headings
        self.NORMAL_FONT = ('Segoe UI', 9)              # Normal text
        self.SMALL_FONT = ('Segoe UI', 8)               # Small text
        self.CALENDAR_MONTH_FONT = ('Segoe UI', 14, 'bold')  # Calendar month display

        # Current state
        self.current_profile_id = None
        self.current_medicine_name = None
        self.current_month_year = None
        self.current_log = None

        # Dirty state tracking for unsaved changes
        self.profile_dirty = False
        self.med_card_dirty = False
        self.log_dirty = False

        # Setup UI
        self._create_menu()
        self._create_main_layout()

        # Load initial data
        self.refresh_profile_list()
    
    def _create_menu(self):
        """Create menu bar"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Exit", command=self.root.quit)

        # Export menu
        export_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Export", menu=export_menu)
        export_menu.add_command(label="Export Medication Log...", command=self.open_export_dialog)

        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self.show_about)
    
    def _create_main_layout(self):
        """Create main application layout with tabs"""
        # Create notebook (tabbed interface)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Tab 1: Profiles
        self.profiles_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.profiles_tab, text="Profiles")
        self._create_profiles_tab()
        
        # Tab 2: Medication Cards
        self.medication_cards_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.medication_cards_tab, text="Medication Cards")
        self._create_medication_cards_tab()

        # Tab 3: Medication Logs
        self.logs_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.logs_tab, text="Medication Logs")
        self._create_logs_tab()

        # Tab 4: Data Editor
        self.editor_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.editor_tab, text="JSON Editor")
        self._create_editor_tab()

        # Tab 5: Settings
        self.settings_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.settings_tab, text="Settings")
        self._create_settings_tab()
    
    def _create_profiles_tab(self):
        """Create profiles management tab"""
        # Use PanedWindow for resizable split
        paned = ttk.PanedWindow(self.profiles_tab, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=(5, 0))

        # Left panel - Profile list
        left_frame = ttk.Frame(paned, width=240)
        paned.add(left_frame, weight=0)

        # Left header with label and button
        left_header = ttk.Frame(left_frame)
        left_header.pack(fill=tk.X, pady=5, padx=5)

        ttk.Label(left_header, text="Patient Profiles", font=self.TITLE_FONT).pack(side=tk.LEFT)
        ttk.Button(left_header, text="New Profile", command=self.new_profile).pack(side=tk.LEFT, padx=10)

        # Profile listbox
        list_frame = ttk.Frame(left_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.profile_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set)
        self.profile_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.profile_listbox.yview)

        self.profile_listbox.bind('<<ListboxSelect>>', self.on_profile_select)

        # Right panel - Profile details
        right_frame = ttk.Frame(paned)
        paned.add(right_frame, weight=1)

        # Right header with label and buttons
        right_header = ttk.Frame(right_frame)
        right_header.pack(fill=tk.X, pady=5, padx=5)

        ttk.Label(right_header, text="Profile Details", font=self.TITLE_FONT).pack(side=tk.LEFT)
        ttk.Button(right_header, text="Delete Profile", command=self.delete_profile).pack(side=tk.RIGHT, padx=2)
        ttk.Button(right_header, text="Save Profile", command=self.save_profile).pack(side=tk.RIGHT, padx=2)

        # Form fields
        form_frame = ttk.Frame(right_frame)
        form_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.profile_fields = {}

        fields = [
            ("child_name", "Child's Name *"),
            ("foster_home", "Foster Home *"),
            ("allergies", "Allergies/Contraindications"),
            ("prescriber_name", "Prescriber Name"),
            ("prescriber_phone", "Prescriber Phone"),
            ("pharmacy", "Pharmacy"),
            ("pharmacy_phone", "Pharmacy Phone")
        ]

        for i, (field_name, label) in enumerate(fields):
            ttk.Label(form_frame, text=label + ":").grid(row=i, column=0, sticky=tk.W, pady=5)
            entry = ttk.Entry(form_frame, width=40)
            entry.grid(row=i, column=1, sticky=tk.EW, pady=5, padx=5)
            entry.bind('<KeyRelease>', self.mark_profile_dirty)
            self.profile_fields[field_name] = entry

        form_frame.columnconfigure(1, weight=1)

        # Status label
        self.profile_status = ttk.Label(right_frame, text="", foreground="green")
        self.profile_status.pack(pady=10)

        # Data location info at bottom of tab
        data_location_frame = ttk.Frame(self.profiles_tab)
        data_location_frame.pack(fill=tk.X, padx=10, pady=5, side=tk.BOTTOM)

        # Get data directory from profile manager
        data_dir = self.profile_manager.data_dir

        # Label showing data location
        ttk.Label(data_location_frame, text="Data Location:",
                 font=self.NORMAL_FONT).pack(side=tk.LEFT, padx=5)

        data_path_label = ttk.Label(data_location_frame, text=data_dir,
                                    font=self.NORMAL_FONT, foreground='blue')
        data_path_label.pack(side=tk.LEFT, padx=5)

        # Open folder button
        ttk.Button(data_location_frame, text="Open Folder",
                  command=self.open_data_folder).pack(side=tk.LEFT, padx=5)

    def _create_medication_cards_tab(self):
        """Create medication cards management tab"""
        # Top panel - Selection
        top_frame = ttk.Frame(self.medication_cards_tab)
        top_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(top_frame, text="Patient:").pack(side=tk.LEFT, padx=5)
        self.med_card_profile_combo = ttk.Combobox(top_frame, width=30, state='readonly')
        self.med_card_profile_combo.pack(side=tk.LEFT, padx=5)
        self.med_card_profile_combo.bind('<<ComboboxSelected>>', self.on_med_card_profile_select)

        ttk.Button(top_frame, text="New Medication Card", command=self.create_new_medication_card).pack(side=tk.LEFT, padx=5)

        # Action buttons on the right
        ttk.Button(top_frame, text="Delete Card", command=self.delete_medication_card).pack(side=tk.RIGHT, padx=5)
        ttk.Button(top_frame, text="Copy Card", command=self.copy_medication_card).pack(side=tk.RIGHT, padx=5)
        ttk.Button(top_frame, text="Save Card", command=self.save_medication_card).pack(side=tk.RIGHT, padx=5)

        # Use PanedWindow for resizable split
        paned = ttk.PanedWindow(self.medication_cards_tab, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Left - Medication card list
        left_frame = ttk.LabelFrame(paned, text="Medication Cards")
        paned.add(left_frame, weight=1)

        self.med_cards_listbox = tk.Listbox(left_frame, height=20)
        self.med_cards_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.med_cards_listbox.bind('<<ListboxSelect>>', self.on_med_card_select)

        # Right - Card details
        right_frame = ttk.LabelFrame(paned, text="Card Details")
        paned.add(right_frame, weight=3)

        # Form fields
        form_frame = ttk.Frame(right_frame)
        form_frame.pack(fill=tk.X, padx=10, pady=5)

        self.med_card_fields = {}

        card_fields = [
            ("medicine_name", "Medicine Name *"),
            ("strength", "Strength"),
            ("dosage", "Dosage"),
            ("starting_amount", "Starting Amount")
        ]

        for i, (field_name, label) in enumerate(card_fields):
            ttk.Label(form_frame, text=label + ":").grid(row=i, column=0, sticky=tk.W, pady=5)
            entry = ttk.Entry(form_frame, width=40)
            entry.grid(row=i, column=1, sticky=tk.EW, pady=5, padx=5)
            entry.bind('<KeyRelease>', self.mark_med_card_dirty)
            self.med_card_fields[field_name] = entry

        form_frame.columnconfigure(1, weight=1)

        # Reason section
        reason_frame = ttk.Frame(right_frame)
        reason_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.med_card_reason_type = tk.StringVar(value="prescribed")

        ttk.Radiobutton(reason_frame, text="Reason Prescribed", variable=self.med_card_reason_type,
                        value="prescribed", command=self.on_med_card_reason_type_change).pack(anchor=tk.W)
        ttk.Radiobutton(reason_frame, text="Reason PRN", variable=self.med_card_reason_type,
                        value="prn", command=self.on_med_card_reason_type_change).pack(anchor=tk.W)

        self.med_card_reason_text = tk.Text(reason_frame, wrap=tk.WORD, height=3, width=50)
        self.med_card_reason_text.pack(fill=tk.BOTH, expand=True, pady=5)
        self.med_card_reason_text.bind('<KeyRelease>', self.mark_med_card_dirty)

        # Images section
        images_frame = ttk.LabelFrame(right_frame, text="Images")
        images_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        img_btn_frame = ttk.Frame(images_frame)
        img_btn_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Button(img_btn_frame, text="Add Image", command=self.add_medication_image).pack(side=tk.LEFT, padx=2)
        ttk.Button(img_btn_frame, text="View Image", command=self.view_medication_image).pack(side=tk.LEFT, padx=2)
        ttk.Button(img_btn_frame, text="Remove Image", command=self.remove_medication_image).pack(side=tk.LEFT, padx=2)

        self.med_card_images_listbox = tk.Listbox(images_frame, height=5)
        self.med_card_images_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def _create_logs_tab(self):
        """Create medication logs tab"""
        # Top panel - Selection
        top_frame = ttk.Frame(self.logs_tab)
        top_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(top_frame, text="Patient:").pack(side=tk.LEFT, padx=5)
        self.log_profile_combo = ttk.Combobox(top_frame, width=30, state='readonly')
        self.log_profile_combo.pack(side=tk.LEFT, padx=5)
        self.log_profile_combo.bind('<<ComboboxSelected>>', self.on_log_profile_select)
        
        ttk.Label(top_frame, text="Log (Medicine - Month/Year):").pack(side=tk.LEFT, padx=5)
        self.log_combo = ttk.Combobox(top_frame, width=40, state='readonly')
        self.log_combo.pack(side=tk.LEFT, padx=5)
        self.log_combo.bind('<<ComboboxSelected>>', self.on_log_select)
        
        ttk.Button(top_frame, text="New Log", command=self.create_new_log).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="Delete Log", command=self.delete_current_log).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="Export", command=self.export_current_log).pack(side=tk.LEFT, padx=5)
        
        # Middle panel - Log details
        details_frame = ttk.LabelFrame(self.logs_tab, text="Medication Information")
        details_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.log_fields = {}
        
        med_fields = [
            ("medicine_name", "Medicine Name"),
            ("strength", "Strength"),
            ("dosage", "Dosage")
        ]
        
        for i, (field_name, label) in enumerate(med_fields):
            ttk.Label(details_frame, text=label + ":").grid(row=0, column=i*2, sticky=tk.W, padx=5, pady=5)
            entry = ttk.Entry(details_frame, width=20)
            entry.grid(row=0, column=i*2+1, sticky=tk.EW, padx=5, pady=5)
            entry.bind('<KeyRelease>', self.mark_log_dirty)
            self.log_fields[field_name] = entry
        
        ttk.Button(details_frame, text="Save Med Info", command=self.save_medication_info).grid(
            row=0, column=6, padx=5, pady=5
        )

        # Reason section
        reason_frame = ttk.LabelFrame(self.logs_tab, text="Reason")
        reason_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Radio button toggle
        radio_frame = ttk.Frame(reason_frame)
        radio_frame.pack(fill=tk.X, padx=5, pady=5)

        self.reason_type = tk.StringVar(value="prescribed")

        ttk.Radiobutton(radio_frame, text="Reason Prescribed", variable=self.reason_type,
                        value="prescribed", command=self.on_reason_type_change).pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(radio_frame, text="Reason PRN", variable=self.reason_type,
                        value="prn", command=self.on_reason_type_change).pack(side=tk.LEFT, padx=5)

        # Resizable text box
        text_frame = ttk.Frame(reason_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.reason_text = tk.Text(text_frame, wrap=tk.WORD, height=4, width=80)
        reason_scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.reason_text.yview)
        self.reason_text.configure(yscrollcommand=reason_scrollbar.set)

        self.reason_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        reason_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.reason_text.bind('<KeyRelease>', self.mark_log_dirty)

        # Save button for reason
        ttk.Button(reason_frame, text="Save Reason", command=self.save_reason).pack(pady=5)

        # Bottom panel - Administration entries
        entry_frame = ttk.LabelFrame(self.logs_tab, text="Administration Entries")
        entry_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Entry form
        input_frame = ttk.Frame(entry_frame)
        input_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(input_frame, text="Day:").grid(row=0, column=0, padx=2)
        self.entry_day = ttk.Spinbox(input_frame, from_=1, to=31, width=5)
        self.entry_day.grid(row=0, column=1, padx=2)
        
        ttk.Label(input_frame, text="Time:").grid(row=0, column=2, padx=2)
        self.entry_time = ttk.Entry(input_frame, width=10)
        self.entry_time.grid(row=0, column=3, padx=2)
        
        ttk.Label(input_frame, text="Initials:").grid(row=0, column=4, padx=2)
        self.entry_initials = ttk.Entry(input_frame, width=5)
        self.entry_initials.grid(row=0, column=5, padx=2)
        
        ttk.Label(input_frame, text="Amount Remaining:").grid(row=0, column=6, padx=2)
        self.entry_amount = ttk.Entry(input_frame, width=15)
        self.entry_amount.grid(row=0, column=7, padx=2)
        
        ttk.Button(input_frame, text="Add Entry", command=self.add_administration_entry).grid(
            row=0, column=8, padx=5
        )
        
        # View toggle controls
        view_control_frame = ttk.Frame(entry_frame)
        view_control_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(view_control_frame, text="View:").pack(side=tk.LEFT, padx=5)

        self.view_mode = tk.StringVar(value="list")
        ttk.Radiobutton(view_control_frame, text="List", variable=self.view_mode,
                       value="list", command=self.toggle_view_mode).pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(view_control_frame, text="Calendar", variable=self.view_mode,
                       value="calendar", command=self.toggle_view_mode).pack(side=tk.LEFT, padx=5)

        # Container for both views
        view_container = ttk.Frame(entry_frame)
        view_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Entries list view
        self.list_view_frame = ttk.Frame(view_container)

        # Treeview for entries
        columns = ('Day', 'Time', 'Initials', 'Amount Remaining')
        self.entries_tree = ttk.Treeview(self.list_view_frame, columns=columns, show='headings', height=15)

        for col in columns:
            self.entries_tree.heading(col, text=col)
            self.entries_tree.column(col, width=150)

        scrollbar = ttk.Scrollbar(self.list_view_frame, orient=tk.VERTICAL, command=self.entries_tree.yview)
        self.entries_tree.configure(yscrollcommand=scrollbar.set)

        self.entries_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Calendar view
        self.calendar_view_frame = ttk.Frame(view_container)
        self._create_calendar_grid()

        # Show list view by default
        self.list_view_frame.pack(fill=tk.BOTH, expand=True)
        
        # Entry buttons
        entry_btn_frame = ttk.Frame(entry_frame)
        entry_btn_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(entry_btn_frame, text="Delete Selected", command=self.delete_administration_entry).pack(
            side=tk.LEFT, padx=2
        )
        ttk.Button(entry_btn_frame, text="Clear All for Day", command=self.clear_day_entries).pack(
            side=tk.LEFT, padx=2
        )
    
    def _create_editor_tab(self):
        """Create JSON editor tab"""
        # Top controls
        top_frame = ttk.Frame(self.editor_tab)
        top_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(top_frame, text="View/Edit JSON Data", font=self.TITLE_FONT).pack(side=tk.LEFT, padx=5)

        ttk.Button(top_frame, text="Load Current Log", command=self.load_json_to_editor).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="Save Changes", command=self.save_json_from_editor).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="Refresh", command=self.load_json_to_editor).pack(side=tk.LEFT, padx=5)
        ttk.Button(top_frame, text="Export to Excel", command=self.export_json_to_excel).pack(side=tk.LEFT, padx=5)

        ttk.Label(top_frame, text="Advanced Users Only", font=self.NORMAL_FONT, foreground='#888888').pack(side=tk.RIGHT, padx=5)
        
        # JSON text editor
        self.json_editor = scrolledtext.ScrolledText(self.editor_tab, wrap=tk.WORD, width=80, height=30)
        self.json_editor.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Status
        self.editor_status = ttk.Label(self.editor_tab, text="", foreground="blue")
        self.editor_status.pack(pady=5)
    
    # Profile methods
    def refresh_profile_list(self):
        """Refresh the profile listbox"""
        self.profile_listbox.delete(0, tk.END)

        profiles = self.profile_manager.get_profile_list()
        for profile_id, child_name in profiles:
            self.profile_listbox.insert(tk.END, child_name)

        # Update log profile combo
        self.log_profile_combo['values'] = [name for _, name in profiles]

        # Update medication cards profile combo
        self.refresh_med_card_profile_list()
    
    def on_profile_select(self, event):
        """Handle profile selection"""
        # Check for unsaved changes
        if not self.check_unsaved_profile_changes():
            # User cancelled, revert selection
            if hasattr(self, 'current_profile_id') and self.current_profile_id:
                profiles = self.profile_manager.get_profile_list()
                for i, (pid, _) in enumerate(profiles):
                    if pid == self.current_profile_id:
                        self.profile_listbox.selection_clear(0, tk.END)
                        self.profile_listbox.selection_set(i)
                        break
            return

        selection = self.profile_listbox.curselection()
        if not selection:
            return

        idx = selection[0]
        profiles = self.profile_manager.get_profile_list()

        if idx < len(profiles):
            profile_id, _ = profiles[idx]
            self.current_profile_id = profile_id
            self.load_profile_to_form(profile_id)
    
    def load_profile_to_form(self, profile_id):
        """Load profile data into form fields"""
        profile = self.profile_manager.get_profile(profile_id)

        if profile:
            for field_name, entry in self.profile_fields.items():
                entry.delete(0, tk.END)
                entry.insert(0, profile.get(field_name, ''))
            # Clear dirty flag after loading
            self.profile_dirty = False
    
    def new_profile(self):
        """Clear form for new profile"""
        # Check for unsaved changes
        if not self.check_unsaved_profile_changes():
            return

        self.current_profile_id = None
        for entry in self.profile_fields.values():
            entry.delete(0, tk.END)
        self.profile_status.config(text="Ready for new profile", foreground="blue")
        self.profile_dirty = False
    
    def save_profile(self):
        """Save profile (create or update)"""
        # Get form data
        profile_data = {}
        for field_name, entry in self.profile_fields.items():
            profile_data[field_name] = entry.get().strip()
        
        # Validate required fields
        if not profile_data.get('child_name'):
            messagebox.showerror("Error", "Child's Name is required")
            return
        
        if not profile_data.get('foster_home'):
            messagebox.showerror("Error", "Foster Home is required")
            return
        
        try:
            if self.current_profile_id:
                # Update existing
                self.profile_manager.update_profile(self.current_profile_id, profile_data)
                self.profile_status.config(text="Profile updated successfully", foreground="green")
            else:
                # Create new
                profile_id = self.profile_manager.create_profile(profile_data)
                self.current_profile_id = profile_id
                self.profile_status.config(text="Profile created successfully", foreground="green")

            self.refresh_profile_list()
            self.profile_dirty = False  # Clear dirty flag after successful save

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save profile: {e}")
    
    def delete_profile(self):
        """Delete selected profile"""
        if not self.current_profile_id:
            messagebox.showwarning("Warning", "No profile selected")
            return
        
        profile = self.profile_manager.get_profile(self.current_profile_id)
        child_name = profile.get('child_name', 'Unknown')
        
        if messagebox.askyesno("Confirm Delete", 
                               f"Delete profile for {child_name}?\n\nThis will NOT delete associated logs."):
            self.profile_manager.delete_profile(self.current_profile_id)
            self.current_profile_id = None
            self.refresh_profile_list()
            self.new_profile()
            self.profile_status.config(text="Profile deleted", foreground="orange")

    # Medication Card methods
    def refresh_med_card_profile_list(self):
        """Refresh medication card profile combo"""
        profiles = self.profile_manager.get_profile_list()
        self.med_card_profile_combo['values'] = [name for _, name in profiles]

    def on_med_card_profile_select(self, event):
        """Handle profile selection in medication cards tab"""
        # Check for unsaved changes
        if not self.check_unsaved_med_card_changes():
            # User cancelled, revert selection
            if hasattr(self, 'current_med_card_profile_id'):
                profiles = self.profile_manager.get_profile_list()
                for profile_id, child_name in profiles:
                    if profile_id == self.current_med_card_profile_id:
                        self.med_card_profile_combo.set(child_name)
                        break
            return

        selected_name = self.med_card_profile_combo.get()

        # Find profile ID
        profiles = self.profile_manager.get_profile_list()
        for profile_id, child_name in profiles:
            if child_name == selected_name:
                self.current_med_card_profile_id = profile_id
                self.load_medication_cards(profile_id)
                break

    def load_medication_cards(self, profile_id):
        """Load medication cards for selected profile"""
        self.med_cards_listbox.delete(0, tk.END)
        self.current_med_card = None

        card_names = self.medication_card_manager.list_cards(profile_id)
        for name in card_names:
            self.med_cards_listbox.insert(tk.END, name)

    def on_med_card_select(self, event):
        """Handle medication card selection"""
        # Check for unsaved changes
        if not self.check_unsaved_med_card_changes():
            # User cancelled, revert selection
            if self.current_med_card:
                cards = list(self.medication_card_manager.get_all_cards(self.current_med_card_profile_id).keys())
                current_name = self.current_med_card.get('medicine_name', '')
                if current_name in cards:
                    index = cards.index(current_name)
                    self.med_cards_listbox.selection_clear(0, tk.END)
                    self.med_cards_listbox.selection_set(index)
            return

        selection = self.med_cards_listbox.curselection()
        if not selection:
            return

        idx = selection[0]
        medicine_name = self.med_cards_listbox.get(idx)

        if hasattr(self, 'current_med_card_profile_id'):
            card = self.medication_card_manager.get_card(self.current_med_card_profile_id, medicine_name)
            if card:
                self.current_med_card = card
                self.load_med_card_to_form(card)

    def load_med_card_to_form(self, card):
        """Load medication card data into form"""
        # Load basic fields
        for field_name, entry in self.med_card_fields.items():
            entry.delete(0, tk.END)
            entry.insert(0, card.get(field_name, ''))

        # Intelligently select which reason to show
        reason_prescribed = card.get('reason_prescribed', '').strip()
        reason_prn = card.get('reason_prn', '').strip()

        # If only PRN has content, show that; otherwise default to prescribed
        if reason_prn and not reason_prescribed:
            self.med_card_reason_type.set("prn")
        else:
            self.med_card_reason_type.set("prescribed")

        # Load reason
        self.load_med_card_reason_text(card)

        # Load images
        self.refresh_med_card_images_list(card)

        # Clear dirty flag after loading
        self.med_card_dirty = False

    def load_med_card_reason_text(self, card):
        """Load reason text based on radio button selection"""
        self.med_card_reason_text.delete('1.0', tk.END)

        if self.med_card_reason_type.get() == "prescribed":
            text = card.get('reason_prescribed', '')
        else:
            text = card.get('reason_prn', '')

        self.med_card_reason_text.insert('1.0', text)

    def on_med_card_reason_type_change(self):
        """Handle reason type change in medication card"""
        if self.current_med_card:
            # Save current text
            text = self.med_card_reason_text.get('1.0', tk.END).strip()

            if self.med_card_reason_type.get() == "prescribed":
                self.current_med_card['reason_prn'] = text
            else:
                self.current_med_card['reason_prescribed'] = text

            # Load other reason type
            self.load_med_card_reason_text(self.current_med_card)

    def refresh_med_card_images_list(self, card):
        """Refresh images list"""
        self.med_card_images_listbox.delete(0, tk.END)

        for image_name in card.get('images', []):
            self.med_card_images_listbox.insert(tk.END, image_name)

    def create_new_medication_card(self):
        """Create new medication card"""
        if not hasattr(self, 'current_med_card_profile_id'):
            messagebox.showwarning("Warning", "Please select a patient first")
            return

        # Check for unsaved changes
        if not self.check_unsaved_med_card_changes():
            return

        # Clear form
        for entry in self.med_card_fields.values():
            entry.delete(0, tk.END)

        self.med_card_reason_text.delete('1.0', tk.END)
        self.med_card_images_listbox.delete(0, tk.END)
        self.current_med_card = None
        self.med_card_dirty = False

    def save_medication_card(self):
        """Save medication card (create or update)"""
        if not hasattr(self, 'current_med_card_profile_id'):
            messagebox.showwarning("Warning", "Please select a patient first")
            return

        # Get form data
        card_data = {}
        for field_name, entry in self.med_card_fields.items():
            card_data[field_name] = entry.get().strip()

        if not card_data.get('medicine_name'):
            messagebox.showerror("Error", "Medicine Name is required")
            return

        # Get reason data
        reason_text = self.med_card_reason_text.get('1.0', tk.END).strip()

        if self.med_card_reason_type.get() == "prescribed":
            card_data['reason_prescribed'] = reason_text
            card_data['reason_prn'] = self.current_med_card.get('reason_prn', '') if self.current_med_card else ''
        else:
            card_data['reason_prn'] = reason_text
            card_data['reason_prescribed'] = self.current_med_card.get('reason_prescribed', '') if self.current_med_card else ''

        try:
            if self.current_med_card is None:
                # Create new
                self.medication_card_manager.create_card(self.current_med_card_profile_id, card_data)
                messagebox.showinfo("Success", "Medication card created")
            else:
                # Update existing
                medicine_name = card_data['medicine_name']
                self.medication_card_manager.update_card(self.current_med_card_profile_id, medicine_name, card_data)
                messagebox.showinfo("Success", "Medication card updated")

            self.load_medication_cards(self.current_med_card_profile_id)
            self.med_card_dirty = False  # Clear dirty flag after successful save

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save card: {e}")

    def copy_medication_card(self):
        """Create a copy of the current medication card"""
        if not self.current_med_card:
            messagebox.showwarning("Warning", "No card selected")
            return

        # Get current card data
        original_name = self.current_med_card.get('medicine_name', '')

        # Generate new name with " - Copy" suffix
        base_name = original_name
        new_name = f"{base_name} - Copy"

        # Check if name already exists, and add numbers if needed
        counter = 2
        while self.medication_card_manager.get_card(self.current_med_card_profile_id, new_name):
            new_name = f"{base_name} - Copy {counter}"
            counter += 1

        # Create copy of card data
        card_data = {}
        for field_name in ['medicine_name', 'strength', 'dosage', 'starting_amount', 'reason_prescribed', 'reason_prn']:
            card_data[field_name] = self.current_med_card.get(field_name, '')

        # Update with new name
        card_data['medicine_name'] = new_name

        try:
            # Create the new card
            self.medication_card_manager.create_card(self.current_med_card_profile_id, card_data)

            # Reload the cards list
            self.load_medication_cards(self.current_med_card_profile_id)

            # Select the new card
            cards = list(self.medication_card_manager.get_all_cards(self.current_med_card_profile_id).keys())
            if new_name in cards:
                index = cards.index(new_name)
                self.med_cards_listbox.selection_clear(0, tk.END)
                self.med_cards_listbox.selection_set(index)
                self.med_cards_listbox.see(index)
                self.on_med_card_select(None)

            messagebox.showinfo("Success", f"Card copied as '{new_name}'")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to copy card: {e}")

    def delete_medication_card(self):
        """Delete selected medication card"""
        if not self.current_med_card:
            messagebox.showwarning("Warning", "No card selected")
            return

        medicine_name = self.current_med_card.get('medicine_name', 'Unknown')

        if messagebox.askyesno("Confirm Delete",
                               f"Delete medication card for {medicine_name}?\n\nThis will also delete associated images."):
            try:
                self.medication_card_manager.delete_card(
                    self.current_med_card_profile_id,
                    medicine_name,
                    delete_images=True
                )
                self.current_med_card = None
                self.load_medication_cards(self.current_med_card_profile_id)
                messagebox.showinfo("Success", "Medication card deleted")

            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete card: {e}")

    def add_medication_image(self):
        """Add image to medication card"""
        if not self.current_med_card:
            messagebox.showwarning("Warning", "No card selected. Please save the card first.")
            return

        # File dialog
        filetypes = (
            ('Image files', '*.jpg *.jpeg *.png *.gif *.bmp'),
            ('PDF files', '*.pdf'),
            ('All files', '*.*')
        )

        filename = filedialog.askopenfilename(title="Select Image", filetypes=filetypes)

        if filename:
            try:
                medicine_name = self.current_med_card.get('medicine_name')
                stored_filename = self.medication_card_manager.add_image(
                    self.current_med_card_profile_id,
                    medicine_name,
                    filename
                )

                # Reload card
                self.current_med_card = self.medication_card_manager.get_card(
                    self.current_med_card_profile_id,
                    medicine_name
                )
                self.refresh_med_card_images_list(self.current_med_card)

                messagebox.showinfo("Success", f"Image '{stored_filename}' added")

            except Exception as e:
                messagebox.showerror("Error", f"Failed to add image: {e}")

    def view_medication_image(self):
        """View selected medication image"""
        if not self.current_med_card:
            messagebox.showwarning("Warning", "No card selected")
            return

        selection = self.med_card_images_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "No image selected")
            return

        idx = selection[0]
        filename = self.med_card_images_listbox.get(idx)

        try:
            medicine_name = self.current_med_card.get('medicine_name')
            image_path = self.medication_card_manager.get_image_path(
                self.current_med_card_profile_id,
                medicine_name,
                filename
            )

            # Open image in default application
            if sys.platform == 'win32':
                os.startfile(image_path)
            elif sys.platform == 'darwin':
                os.system(f'open "{image_path}"')
            else:
                os.system(f'xdg-open "{image_path}"')

        except Exception as e:
            messagebox.showerror("Error", f"Failed to view image: {e}")

    def remove_medication_image(self):
        """Remove selected medication image"""
        if not self.current_med_card:
            messagebox.showwarning("Warning", "No card selected")
            return

        selection = self.med_card_images_listbox.curselection()
        if not selection:
            messagebox.showwarning("Warning", "No image selected")
            return

        idx = selection[0]
        filename = self.med_card_images_listbox.get(idx)

        if messagebox.askyesno("Confirm", f"Remove image '{filename}'?"):
            try:
                medicine_name = self.current_med_card.get('medicine_name')
                self.medication_card_manager.remove_image(
                    self.current_med_card_profile_id,
                    medicine_name,
                    filename
                )

                # Reload card
                self.current_med_card = self.medication_card_manager.get_card(
                    self.current_med_card_profile_id,
                    medicine_name
                )
                self.refresh_med_card_images_list(self.current_med_card)

                messagebox.showinfo("Success", "Image removed")

            except Exception as e:
                messagebox.showerror("Error", f"Failed to remove image: {e}")

    # Log methods
    def on_log_profile_select(self, event):
        """Handle profile selection in logs tab"""
        # Check for unsaved changes
        if not self.check_unsaved_log_changes():
            # User cancelled, revert selection
            if hasattr(self, 'current_profile_id') and self.current_profile_id:
                profiles = self.profile_manager.get_profile_list()
                for profile_id, child_name in profiles:
                    if profile_id == self.current_profile_id:
                        self.log_profile_combo.set(child_name)
                        break
            return

        selected_name = self.log_profile_combo.get()

        # Find profile ID
        profiles = self.profile_manager.get_profile_list()
        for profile_id, child_name in profiles:
            if child_name == selected_name:
                self.current_profile_id = profile_id
                self.load_logs_for_profile(profile_id)
                break

    def load_logs_for_profile(self, profile_id):
        """Load available logs for selected profile"""
        logs = self.log_manager.list_logs_for_profile(profile_id)  # Returns list of (medicine_name, month_year) tuples

        # Format for display: "Medicine Name - Month Year"
        display_logs = [f"{med} - {month}" for med, month in logs]
        self.log_combo['values'] = display_logs

        # Store the tuples for later reference
        self.available_logs = logs

        if logs:
            self.log_combo.current(0)
            self.on_log_select(None)

    def on_log_select(self, event):
        """Handle log selection"""
        # Check for unsaved changes
        if not self.check_unsaved_log_changes():
            # User cancelled, revert selection
            if hasattr(self, 'current_medicine_name') and self.current_medicine_name and hasattr(self, 'current_month_year') and self.current_month_year:
                for i, (med, month) in enumerate(self.available_logs):
                    if med == self.current_medicine_name and month == self.current_month_year:
                        self.log_combo.current(i)
                        break
            return

        idx = self.log_combo.current()

        if idx >= 0 and idx < len(self.available_logs):
            self.current_medicine_name, self.current_month_year = self.available_logs[idx]
            self.load_current_log()
    
    def create_new_log(self):
        """Create new medication log"""
        if not self.current_profile_id:
            messagebox.showwarning("Warning", "Please select a patient first")
            return

        # Check for unsaved changes
        if not self.check_unsaved_log_changes():
            return

        # Dialog for new log
        dialog = tk.Toplevel(self.root)
        dialog.title("New Medication Log")
        dialog.geometry("500x350")

        # Option to select from medication card or enter manually
        ttk.Label(dialog, text="Select from Medication Card:").grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)

        # Get medication cards for this profile
        card_names = self.medication_card_manager.list_cards(self.current_profile_id)
        card_names.insert(0, "-- Enter Manually --")

        card_combo = ttk.Combobox(dialog, width=25, values=card_names, state='readonly')
        card_combo.grid(row=0, column=1, columnspan=3, padx=10, pady=10, sticky=tk.W)
        card_combo.current(0)

        ttk.Label(dialog, text="Medicine Name:").grid(row=1, column=0, padx=10, pady=10, sticky=tk.W)
        med_entry = ttk.Entry(dialog, width=25)
        med_entry.grid(row=1, column=1, columnspan=3, padx=10, pady=10, sticky=tk.W)

        # Month and Year selection
        ttk.Label(dialog, text="Month:").grid(row=2, column=0, padx=10, pady=10, sticky=tk.W)

        months = ['January', 'February', 'March', 'April', 'May', 'June',
                 'July', 'August', 'September', 'October', 'November', 'December']

        month_combo = ttk.Combobox(dialog, width=12, values=months, state='readonly')
        month_combo.grid(row=2, column=1, padx=10, pady=10, sticky=tk.W)
        month_combo.current(datetime.now().month - 1)  # Set to current month

        ttk.Label(dialog, text="Year:").grid(row=2, column=2, padx=(5, 5), sticky=tk.W)
        year_entry = ttk.Entry(dialog, width=6)
        year_entry.grid(row=2, column=3, padx=(0, 10), pady=10, sticky=tk.W)
        year_entry.insert(0, str(datetime.now().year))  # Set to current year

        ttk.Label(dialog, text="Strength:").grid(row=3, column=0, padx=10, pady=10, sticky=tk.W)
        strength_entry = ttk.Entry(dialog, width=25)
        strength_entry.grid(row=3, column=1, columnspan=3, padx=10, pady=10, sticky=tk.W)

        ttk.Label(dialog, text="Dosage:").grid(row=4, column=0, padx=10, pady=10, sticky=tk.W)
        dosage_entry = ttk.Entry(dialog, width=25)
        dosage_entry.grid(row=4, column=1, columnspan=3, padx=10, pady=10, sticky=tk.W)

        def on_card_select(event):
            """Handle medication card selection"""
            selected = card_combo.get()

            if selected != "-- Enter Manually --":
                # Load card data
                card = self.medication_card_manager.get_card(self.current_profile_id, selected)

                if card:
                    med_entry.delete(0, tk.END)
                    med_entry.insert(0, card.get('medicine_name', ''))

                    strength_entry.delete(0, tk.END)
                    strength_entry.insert(0, card.get('strength', ''))

                    dosage_entry.delete(0, tk.END)
                    dosage_entry.insert(0, card.get('dosage', ''))
            else:
                # Clear fields
                med_entry.delete(0, tk.END)
                strength_entry.delete(0, tk.END)
                dosage_entry.delete(0, tk.END)

        card_combo.bind('<<ComboboxSelected>>', on_card_select)

        def create():
            medicine_name = med_entry.get().strip()
            month = month_combo.get().strip()
            year = year_entry.get().strip()

            if not medicine_name:
                messagebox.showerror("Error", "Medicine Name is required")
                return

            if not month:
                messagebox.showerror("Error", "Month is required")
                return

            if not year:
                messagebox.showerror("Error", "Year is required")
                return

            # Combine month and year
            month_year = f"{month} {year}"

            if self.log_manager.log_exists(self.current_profile_id, medicine_name, month_year):
                messagebox.showerror("Error", f"Log for {medicine_name} - {month_year} already exists")
                return

            medication_info = {
                'medicine_name': medicine_name,
                'strength': strength_entry.get().strip(),
                'dosage': dosage_entry.get().strip()
            }

            # If using medication card, copy reason fields
            selected_card = card_combo.get()
            if selected_card != "-- Enter Manually --":
                card = self.medication_card_manager.get_card(self.current_profile_id, selected_card)
                if card:
                    medication_info['reason_prescribed'] = card.get('reason_prescribed', '')
                    medication_info['reason_prn'] = card.get('reason_prn', '')

            self.log_manager.create_log(self.current_profile_id, month_year, medication_info)
            dialog.destroy()

            self.load_logs_for_profile(self.current_profile_id)
            messagebox.showinfo("Success", f"Log for {medicine_name} - {month_year} created")

        ttk.Button(dialog, text="Create", command=create).grid(row=5, column=0, columnspan=2, pady=20)
    
    def load_current_log(self):
        """Load the currently selected log"""
        if not self.current_profile_id or not self.current_medicine_name or not self.current_month_year:
            return

        self.current_log = self.log_manager.get_log(self.current_profile_id, self.current_medicine_name, self.current_month_year)

        if self.current_log:
            # Load medication info
            for field_name, entry in self.log_fields.items():
                entry.delete(0, tk.END)
                entry.insert(0, self.current_log.get(field_name, ''))

            # Load reason based on current selection
            self.load_reason_text()

            # Load entries
            self.refresh_entries_list()

            # Clear dirty flag after loading
            self.log_dirty = False
    
    def save_medication_info(self):
        """Save medication information changes"""
        if not self.current_log:
            messagebox.showwarning("Warning", "No log loaded")
            return

        for field_name, entry in self.log_fields.items():
            self.current_log[field_name] = entry.get().strip()

        self.log_manager.save_log(self.current_profile_id, self.current_medicine_name, self.current_month_year, self.current_log)
        messagebox.showinfo("Success", "Medication information saved")
        self.log_dirty = False  # Clear dirty flag after successful save

    def on_reason_type_change(self):
        """Handle reason type radio button change"""
        if self.current_log:
            # Save current text before switching
            self.save_current_reason_to_log()
            # Load the other reason type
            self.load_reason_text()

    def load_reason_text(self):
        """Load the appropriate reason text based on radio button selection"""
        if not self.current_log:
            return

        self.reason_text.delete('1.0', tk.END)

        if self.reason_type.get() == "prescribed":
            text = self.current_log.get('reason_prescribed', '')
        else:
            text = self.current_log.get('reason_prn', '')

        self.reason_text.insert('1.0', text)

    def save_current_reason_to_log(self):
        """Save current reason text to the log data (without persisting to file)"""
        if not self.current_log:
            return

        text = self.reason_text.get('1.0', tk.END).strip()

        if self.reason_type.get() == "prescribed":
            self.current_log['reason_prescribed'] = text
        else:
            self.current_log['reason_prn'] = text

    def save_reason(self):
        """Save reason text to file"""
        if not self.current_log:
            messagebox.showwarning("Warning", "No log loaded")
            return

        # Save current text to log data
        self.save_current_reason_to_log()

        # Persist to file
        self.log_manager.save_log(self.current_profile_id, self.current_medicine_name, self.current_month_year, self.current_log)
        messagebox.showinfo("Success", "Reason saved")
        self.log_dirty = False  # Clear dirty flag after successful save
    
    def refresh_entries_list(self):
        """Refresh the administration entries treeview and calendar view"""
        # Clear existing
        for item in self.entries_tree.get_children():
            self.entries_tree.delete(item)

        if not self.current_log:
            return

        # Add entries
        for entry in self.current_log.get('administration_log', []):
            self.entries_tree.insert('', tk.END, values=(
                entry.get('day', ''),
                entry.get('time', ''),
                entry.get('initials', ''),
                entry.get('amount_remaining', '')
            ))

        # Also refresh calendar view if it's visible
        if self.view_mode.get() == "calendar":
            self.refresh_calendar_view()
    
    def add_administration_entry(self):
        """Add new administration entry"""
        if not self.current_log:
            messagebox.showwarning("Warning", "No log loaded")
            return
        
        try:
            day = int(self.entry_day.get())
            if day < 1 or day > 31:
                raise ValueError("Day must be 1-31")
        except ValueError:
            messagebox.showerror("Error", "Invalid day (must be 1-31)")
            return
        
        entry_data = {
            'day': day,
            'time': self.entry_time.get().strip(),
            'initials': self.entry_initials.get().strip(),
            'amount_remaining': self.entry_amount.get().strip()
        }
        
        if not entry_data['time'] or not entry_data['initials']:
            messagebox.showwarning("Warning", "Time and Initials are required")
            return
        
        try:
            self.log_manager.add_entry(self.current_profile_id, self.current_medicine_name, self.current_month_year, entry_data)
            self.current_log = self.log_manager.get_log(self.current_profile_id, self.current_medicine_name, self.current_month_year)
            self.refresh_entries_list()
            
            # Clear entry fields
            self.entry_time.delete(0, tk.END)
            self.entry_initials.delete(0, tk.END)
            self.entry_amount.delete(0, tk.END)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add entry: {e}")
    
    def delete_administration_entry(self):
        """Delete selected administration entry"""
        selection = self.entries_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "No entry selected")
            return
        
        item = self.entries_tree.item(selection[0])
        values = item['values']
        day = int(values[0])
        
        # Find which entry it is for this day
        day_entries = [e for e in self.current_log.get('administration_log', []) if e['day'] == day]
        
        # Match by all fields
        for idx, entry in enumerate(day_entries):
            if (entry.get('time') == values[1] and 
                entry.get('initials') == values[2] and
                entry.get('amount_remaining') == values[3]):
                
                if messagebox.askyesno("Confirm", f"Delete entry for Day {day}?"):
                    try:
                        self.log_manager.delete_entry(self.current_profile_id, self.current_medicine_name, self.current_month_year, day, idx)
                        self.current_log = self.log_manager.get_log(self.current_profile_id, self.current_medicine_name, self.current_month_year)
                        self.refresh_entries_list()
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to delete entry: {e}")
                break
    
    def clear_day_entries(self):
        """Clear all entries for a specific day"""
        try:
            day = int(self.entry_day.get())
            if day < 1 or day > 31:
                raise ValueError("Day must be 1-31")
        except ValueError:
            messagebox.showerror("Error", "Invalid day (must be 1-31)")
            return
        
        day_entries = [e for e in self.current_log.get('administration_log', []) if e['day'] == day]
        
        if not day_entries:
            messagebox.showinfo("Info", f"No entries for day {day}")
            return
        
        if messagebox.askyesno("Confirm", f"Delete all {len(day_entries)} entries for day {day}?"):
            try:
                self.log_manager.delete_all_entries_for_day(self.current_profile_id, self.current_medicine_name, self.current_month_year, day)
                self.current_log = self.log_manager.get_log(self.current_profile_id, self.current_medicine_name, self.current_month_year)
                self.refresh_entries_list()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to clear entries: {e}")
    
    def delete_current_log(self):
        """Delete the currently loaded log"""
        if not self.current_profile_id or not self.current_medicine_name or not self.current_month_year:
            messagebox.showwarning("Warning", "No log selected")
            return

        if messagebox.askyesno("Confirm Delete",
                               f"Delete log for {self.current_medicine_name} - {self.current_month_year}?\n\nThis cannot be undone."):
            try:
                self.log_manager.delete_log(self.current_profile_id, self.current_medicine_name, self.current_month_year)
                self.current_log = None
                self.current_medicine_name = None
                self.current_month_year = None
                self.load_logs_for_profile(self.current_profile_id)
                messagebox.showinfo("Success", "Log deleted")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete log: {e}")

    # Calendar View methods
    def _create_calendar_grid(self):
        """Create the calendar grid for displaying entries"""
        import calendar

        # Calendar header (month/year)
        header_frame = ttk.Frame(self.calendar_view_frame)
        header_frame.pack(fill=tk.X, padx=5, pady=5)

        self.calendar_month_label = ttk.Label(header_frame, text="", font=self.CALENDAR_MONTH_FONT)
        self.calendar_month_label.pack()

        # Create scrollable canvas for calendar
        canvas_frame = ttk.Frame(self.calendar_view_frame)
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        canvas = tk.Canvas(canvas_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=canvas.yview)

        # Frame to hold the calendar grid
        days_frame = ttk.Frame(canvas)

        # Configure canvas scrolling
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas_window = canvas.create_window((0, 0), window=days_frame, anchor=tk.NW)

        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Update canvas window width to match canvas width
            canvas.itemconfig(canvas_window, width=event.width)

        def on_canvas_configure(event):
            # Update canvas window width when canvas is resized
            canvas.itemconfig(canvas_window, width=event.width)

        days_frame.bind('<Configure>', on_frame_configure)
        canvas.bind('<Configure>', on_canvas_configure)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Enable mousewheel scrolling
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        # Store handler for cleanup
        self.calendar_mousewheel_handler = on_mousewheel
        canvas.bind_all("<MouseWheel>", on_mousewheel)

        # Day headers
        day_headers = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

        # Configure grid columns for equal sizing
        for i in range(7):
            days_frame.columnconfigure(i, weight=1, minsize=80)

        # Day headers
        for col, day in enumerate(day_headers):
            header = ttk.Label(days_frame, text=day, font=self.HEADING_FONT,
                             anchor=tk.CENTER, relief=tk.RIDGE)
            header.grid(row=0, column=col, sticky='ew', padx=1, pady=1)

        # Calendar cells storage
        self.calendar_cells = {}

        # Store canvas reference for cleanup
        self.calendar_canvas = canvas
        self.calendar_mousewheel_binding = None

        # Create 5 rows of 7 cells for the calendar
        for row in range(1, 6):
            for col in range(7):
                cell_frame = tk.Frame(days_frame, relief=tk.RIDGE, borderwidth=1,
                                    bg='white', cursor='hand2', height=60)
                cell_frame.grid(row=row, column=col, sticky='ew', padx=1, pady=1)
                cell_frame.grid_propagate(False)  # Don't shrink to fit content

                # Day number label
                day_label = tk.Label(cell_frame, text="", font=self.HEADING_FONT,
                                   bg='white', anchor=tk.NW)
                day_label.pack(anchor=tk.NW, padx=2, pady=2)

                # Entries container with wrapping
                entries_label = tk.Label(cell_frame, text="", font=self.SMALL_FONT,
                                       bg='white', anchor=tk.NW, justify=tk.LEFT,
                                       wraplength=70)
                entries_label.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

                # Store references including event handler
                self.calendar_cells[(row-1, col)] = {
                    'frame': cell_frame,
                    'day_label': day_label,
                    'entries_label': entries_label,
                    'day_num': 0,
                    'click_handler': None  # Store click handler for cleanup
                }

    def toggle_view_mode(self):
        """Toggle between list and calendar view"""
        if self.view_mode.get() == "list":
            # Cleanup calendar view bindings when switching away
            self._cleanup_calendar_bindings()
            self.calendar_view_frame.pack_forget()
            self.list_view_frame.pack(fill=tk.BOTH, expand=True)
        else:
            self.list_view_frame.pack_forget()
            self.calendar_view_frame.pack(fill=tk.BOTH, expand=True)
            self.refresh_calendar_view()

    def _cleanup_calendar_bindings(self):
        """Cleanup calendar view event bindings to prevent memory leaks"""
        # Unbind mousewheel handler
        if hasattr(self, 'calendar_mousewheel_handler') and self.calendar_mousewheel_handler:
            try:
                self.root.unbind_all("<MouseWheel>")
            except:
                pass  # Best effort cleanup

        # Unbind all cell click handlers
        if hasattr(self, 'calendar_cells'):
            for cell_key, cell in self.calendar_cells.items():
                if cell.get('click_handler'):
                    try:
                        cell['frame'].unbind('<Button-1>')
                        cell['day_label'].unbind('<Button-1>')
                        cell['entries_label'].unbind('<Button-1>')
                        cell['click_handler'] = None
                    except:
                        pass  # Widget may have been destroyed

    def refresh_calendar_view(self):
        """Refresh the calendar view with current log data"""
        import calendar
        from datetime import datetime

        if not self.current_log:
            return

        # Parse month and year from log
        month_year = self.current_log.get('month_year', '')
        try:
            # Parse "December 2025" format
            date_obj = datetime.strptime(month_year, "%B %Y")
            month = date_obj.month
            year = date_obj.year
        except:
            return

        # Update header
        self.calendar_month_label.config(text=month_year)

        # Get calendar for this month
        cal = calendar.monthcalendar(year, month)

        # Group entries by day
        entries_by_day = {}
        for entry in self.current_log.get('administration_log', []):
            day = entry.get('day', 0)
            if day not in entries_by_day:
                entries_by_day[day] = []
            entries_by_day[day].append(entry)

        # Populate calendar cells
        cell_idx = 0

        for week_idx, week in enumerate(cal):
            for day_idx, day in enumerate(week):
                if cell_idx >= len(self.calendar_cells):
                    break

                cell = self.calendar_cells[(week_idx, day_idx)]
                frame = cell['frame']
                day_label = cell['day_label']
                entries_label = cell['entries_label']

                # Unbind old click handlers to prevent memory leak
                if cell.get('click_handler'):
                    frame.unbind('<Button-1>')
                    day_label.unbind('<Button-1>')
                    entries_label.unbind('<Button-1>')
                    cell['click_handler'] = None

                if day == 0:
                    # Empty cell (day from prev/next month)
                    day_label.config(text="")
                    entries_label.config(text="")
                    frame.config(bg='#f0f0f0')
                    day_label.config(bg='#f0f0f0')
                    entries_label.config(bg='#f0f0f0')
                    cell['day_num'] = 0
                else:
                    # Valid day
                    day_label.config(text=str(day))
                    cell['day_num'] = day

                    # Get entries for this day
                    day_entries = entries_by_day.get(day, [])

                    if day_entries:
                        # Has entries - highlight
                        frame.config(bg='#e8f4f8')
                        day_label.config(bg='#e8f4f8')
                        entries_label.config(bg='#e8f4f8')

                        # Show detailed info (always)
                        entry_text = '\n'.join([
                            f"{e.get('time', '')} {e.get('initials', '')}"
                            for e in day_entries[:3]  # Show max 3
                        ])
                        if len(day_entries) > 3:
                            entry_text += f"\n+{len(day_entries)-3} more"
                        entries_label.config(text=entry_text, font=self.SMALL_FONT)
                    else:
                        # No entries
                        frame.config(bg='white')
                        day_label.config(bg='white')
                        entries_label.config(bg='white')
                        entries_label.config(text="", font=self.SMALL_FONT)

                    # Bind click to add entry for this day
                    def make_click_handler(day_num):
                        return lambda e: self.on_calendar_day_click(day_num)

                    handler = make_click_handler(day)
                    frame.bind('<Button-1>', handler)
                    day_label.bind('<Button-1>', handler)
                    entries_label.bind('<Button-1>', handler)

                    # Store handler reference for cleanup
                    cell['click_handler'] = handler

    def on_calendar_day_click(self, day):
        """Handle click on calendar day - set day in entry form"""
        self.entry_day.delete(0, tk.END)
        self.entry_day.insert(0, str(day))
        # Switch to list view to show the entry form is ready
        self.view_mode.set("list")
        self.toggle_view_mode()
        # Focus on time field
        self.entry_time.focus()

    # JSON Editor methods
    def load_json_to_editor(self):
        """Load current log data to JSON editor"""
        if not self.current_log:
            self.editor_status.config(text="No log loaded. Go to Medication Logs tab first.", foreground="orange")
            return
        
        # Format JSON nicely
        json_text = json.dumps(self.current_log, indent=2)
        
        self.json_editor.delete('1.0', tk.END)
        self.json_editor.insert('1.0', json_text)
        
        self.editor_status.config(text="Loaded current log", foreground="green")
    
    def save_json_from_editor(self):
        """Save edited JSON back to log"""
        if not self.current_profile_id or not self.current_medicine_name or not self.current_month_year:
            messagebox.showwarning("Warning", "No log loaded")
            return

        try:
            # Parse JSON
            json_text = self.json_editor.get('1.0', tk.END)
            log_data = json.loads(json_text)

            # Validate it has required fields
            if 'administration_log' not in log_data:
                raise ValueError("Missing 'administration_log' field")

            # Save
            self.log_manager.save_log(self.current_profile_id, self.current_medicine_name, self.current_month_year, log_data)
            self.current_log = log_data
            self.refresh_entries_list()

            self.editor_status.config(text="Changes saved successfully", foreground="green")
            messagebox.showinfo("Success", "JSON changes saved")

        except json.JSONDecodeError as e:
            messagebox.showerror("JSON Error", f"Invalid JSON: {e}")
            self.editor_status.config(text="Invalid JSON - not saved", foreground="red")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {e}")
            self.editor_status.config(text="Error saving changes", foreground="red")

    def _export_log_to_excel(self, log_data, filename):
        """Helper method to export log data to Excel file"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Medication Log"

            # Add header
            ws['A1'] = "Medication Log Data"
            ws['A1'].font = Font(size=16, bold=True)

            # Add basic info
            row = 3
            ws[f'A{row}'] = "Medicine Name:"
            ws[f'B{row}'] = log_data.get('medicine_name', '')
            row += 1

            ws[f'A{row}'] = "Strength:"
            ws[f'B{row}'] = log_data.get('strength', '')
            row += 1

            ws[f'A{row}'] = "Dosage:"
            ws[f'B{row}'] = log_data.get('dosage', '')
            row += 1

            ws[f'A{row}'] = "Month/Year:"
            ws[f'B{row}'] = log_data.get('month_year', '')
            row += 1

            ws[f'A{row}'] = "Reason Prescribed:"
            ws[f'B{row}'] = log_data.get('reason_prescribed', '')
            row += 1

            ws[f'A{row}'] = "Reason PRN:"
            ws[f'B{row}'] = log_data.get('reason_prn', '')
            row += 2

            # Add administration log table
            ws[f'A{row}'] = "Administration Log"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 1

            # Table headers
            headers = ['Day', 'Time', 'Initials', 'Amount Remaining']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            row += 1

            # Add administration entries
            for entry in log_data.get('administration_log', []):
                ws.cell(row=row, column=1, value=entry.get('day', ''))
                ws.cell(row=row, column=2, value=entry.get('time', ''))
                ws.cell(row=row, column=3, value=entry.get('initials', ''))
                ws.cell(row=row, column=4, value=entry.get('amount_remaining', ''))
                row += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save workbook
            wb.save(filename)

        except ImportError:
            raise ImportError("openpyxl library not installed. Install it with: pip install openpyxl")
        except Exception as e:
            raise Exception(f"Failed to export to Excel: {e}")

    def export_json_to_excel(self):
        """Export current JSON data to Excel from JSON editor"""
        if not self.current_log:
            messagebox.showwarning("Warning", "No log loaded to export")
            return

        # Ask user for save location
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Excel File"
        )

        if not filename:
            return

        try:
            self._export_log_to_excel(self.current_log, filename)
            messagebox.showinfo("Success", f"Exported to Excel:\n{filename}")
            self.editor_status.config(text=f"Exported to {os.path.basename(filename)}", foreground="green")
        except ImportError as e:
            messagebox.showerror("Error", str(e))
            self.editor_status.config(text="openpyxl not installed", foreground="red")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.editor_status.config(text="Export failed", foreground="red")
    
    # Export methods
    def export_current_log(self):
        """Export current log to Word/PDF"""
        if not self.current_log or not self.current_profile_id:
            messagebox.showwarning("Warning", "No log loaded to export")
            return

        # Get profile data
        profile = self.profile_manager.get_profile(self.current_profile_id)

        # Create export options dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Export Options")
        dialog.geometry("450x280")
        dialog.minsize(400, 280)

        ttk.Label(dialog, text="Select Export Method:", font=self.HEADING_FONT).pack(pady=10)

        export_method = tk.StringVar(value="extended_table")

        # Radio buttons for export method
        radio_frame = ttk.Frame(dialog)
        radio_frame.pack(pady=10)

        ttk.Radiobutton(
            radio_frame,
            text="Extended Table (adds rows to same page)",
            variable=export_method,
            value="extended_table"
        ).pack(anchor=tk.W, padx=20, pady=5)

        ttk.Radiobutton(
            radio_frame,
            text="Continuation Pages (creates new pages)",
            variable=export_method,
            value="continuation_pages"
        ).pack(anchor=tk.W, padx=20, pady=5)

        # Checkbox for including images
        include_images = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            dialog,
            text="Include medication card images (Word only)",
            variable=include_images
        ).pack(pady=5)

        # Export format selection
        ttk.Label(dialog, text="Export Formats:", font=self.HEADING_FONT).pack(pady=(10, 5))

        format_frame = ttk.Frame(dialog)
        format_frame.pack(pady=5)

        export_word = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            format_frame,
            text="Word (.docx)",
            variable=export_word
        ).pack(side=tk.LEFT, padx=10)

        export_pdf = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            format_frame,
            text="PDF (.pdf)",
            variable=export_pdf
        ).pack(side=tk.LEFT, padx=10)

        export_excel = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            format_frame,
            text="Excel (.xlsx)",
            variable=export_excel
        ).pack(side=tk.LEFT, padx=10)

        # Open folder after export checkbox
        open_folder_after = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            dialog,
            text="Open folder after export",
            variable=open_folder_after
        ).pack(pady=10, padx=20, anchor=tk.W)

        def proceed_with_export():
            # Validate at least one format is selected
            if not export_word.get() and not export_pdf.get() and not export_excel.get():
                messagebox.showwarning("No Format Selected", "Please select at least one export format.")
                return

            dialog.destroy()

            # Ask user if they want to use default patient folder or custom
            use_default = messagebox.askyesno(
                "Export Location",
                "Use patient's default export folder?\n\n"
                "Yes: Save to patient folder (auto-organized)\n"
                "No: Choose custom location"
            )

            output_dir = None
            if not use_default:
                # Use default export folder from settings as initial directory
                initial_dir = self.settings_manager.get_default_export_folder() if self.settings_manager else str(Path.home() / "Documents")
                output_dir = filedialog.askdirectory(
                    title="Select Output Directory",
                    initialdir=initial_dir
                )
                if not output_dir:
                    return

            result = {}
            msg = "Exported successfully:\n\n"

            try:
                # Export Word/PDF if requested
                if export_word.get() or export_pdf.get():
                    word_pdf_result = self.export_manager.export_log(
                        profile,
                        self.current_log,
                        output_dir=output_dir,
                        create_pdf=export_pdf.get(),
                        export_method=export_method.get(),
                        include_images=include_images.get(),
                        data_dir=os.path.join(self.data_dir, "patients")
                    )
                    result.update(word_pdf_result)

                    if 'docx' in result:
                        msg += f"{result['docx']}\n"
                    if 'pdf' in result:
                        msg += f"{result['pdf']}\n"

                # Export to Excel if requested
                if export_excel.get():
                    # Determine Excel filename
                    if 'docx' in result:
                        excel_path = result['docx'].replace('.docx', '.xlsx')
                    else:
                        # Excel-only export, need to generate filename
                        if output_dir is None:
                            profile_id = profile.get('child_name', '').lower().replace(' ', '_')
                            output_dir = self.export_manager.get_patient_export_dir(profile_id, os.path.join(self.data_dir, "patients"))

                        medicine_name = self.current_log.get('medicine_name', 'unknown').replace(' ', '_')
                        month_year = self.current_log['month_year'].replace(' ', '_')
                        filename_base = f"{medicine_name}_{month_year}"
                        excel_path = os.path.join(output_dir, f"{filename_base}.xlsx")

                    try:
                        self._export_log_to_excel(self.current_log, excel_path)
                        result['xlsx'] = excel_path
                        msg += f"{excel_path}\n"
                    except Exception as excel_error:
                        if export_word.get() or export_pdf.get():
                            messagebox.showwarning("Excel Export", f"Word/PDF exported successfully, but Excel export failed: {excel_error}")
                        else:
                            raise

                messagebox.showinfo("Export Complete", msg.strip())

                # Open folder if requested
                if open_folder_after.get() and result:
                    # Get the directory from the first exported file
                    export_dir = None
                    for key in ['docx', 'pdf', 'xlsx']:
                        if key in result:
                            export_dir = os.path.dirname(result[key])
                            break

                    if export_dir:
                        try:
                            system = platform.system()
                            if system == 'Windows':
                                os.startfile(export_dir)
                            elif system == 'Darwin':
                                subprocess.run(['open', export_dir], check=True)
                            else:
                                subprocess.run(['xdg-open', export_dir], check=True)
                        except Exception as folder_error:
                            # Silently fail - folder opening is a convenience feature
                            pass

            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export: {e}")

        # Buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=15, padx=10)

        ttk.Button(btn_frame, text="Export", command=proceed_with_export, padding=5).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy, padding=5).pack(side=tk.LEFT, padx=5)

    def open_export_dialog(self):
        """Open enhanced export dialog with patient and log selection"""
        # Create main export dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Export Medication Log")
        dialog.geometry("500x650")
        dialog.minsize(480, 600)

        # Patient Selection Section
        ttk.Label(dialog, text="1. Select Patient:", font=self.HEADING_FONT).pack(pady=(10, 5))

        patient_frame = ttk.Frame(dialog)
        patient_frame.pack(fill=tk.BOTH, padx=20, pady=5)

        patient_listbox = tk.Listbox(patient_frame, height=5, exportselection=False)
        patient_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        patient_scrollbar = ttk.Scrollbar(patient_frame, orient=tk.VERTICAL, command=patient_listbox.yview)
        patient_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        patient_listbox.config(yscrollcommand=patient_scrollbar.set)

        # Load patients
        profiles = self.profile_manager.get_profile_list()
        patient_map = {}  # Maps listbox index to profile_id
        for idx, (profile_id, profile_name) in enumerate(profiles):
            patient_listbox.insert(tk.END, profile_name)
            patient_map[idx] = profile_id

        # Medication Log Selection Section
        ttk.Label(dialog, text="2. Select Medication Log(s):", font=self.HEADING_FONT).pack(pady=(15, 5))
        ttk.Label(dialog, text="(Ctrl+Click to select multiple, Shift+Click for range)",
                  font=self.NORMAL_FONT, foreground='#666666').pack(pady=(0, 5))

        log_frame = ttk.Frame(dialog)
        log_frame.pack(fill=tk.BOTH, padx=20, pady=5)

        log_listbox = tk.Listbox(log_frame, height=6, exportselection=False, selectmode=tk.EXTENDED)
        log_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        log_scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=log_listbox.yview)
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        log_listbox.config(yscrollcommand=log_scrollbar.set)

        log_map = {}  # Maps listbox index to (medicine_name, month_year)

        # Selection helper buttons
        selection_btn_frame = ttk.Frame(dialog)
        selection_btn_frame.pack(pady=5)

        def select_all_logs():
            """Select all logs in the listbox"""
            log_listbox.selection_set(0, tk.END)

        def clear_selection():
            """Clear all log selections"""
            log_listbox.selection_clear(0, tk.END)

        def select_by_month():
            """Select all logs for a specific month/year"""
            if not log_map:
                messagebox.showinfo("No Logs", "No logs available to filter.")
                return

            # Create month selection dialog
            month_dialog = tk.Toplevel(dialog)
            month_dialog.title("Select Month/Year")
            month_dialog.geometry("350x200")
            month_dialog.transient(dialog)
            month_dialog.grab_set()

            ttk.Label(month_dialog, text="Select logs for:", font=self.HEADING_FONT).pack(pady=10)

            select_frame = ttk.Frame(month_dialog)
            select_frame.pack(pady=10)

            ttk.Label(select_frame, text="Month:").grid(row=0, column=0, padx=5, pady=5)
            months = ['January', 'February', 'March', 'April', 'May', 'June',
                     'July', 'August', 'September', 'October', 'November', 'December']
            month_combo = ttk.Combobox(select_frame, values=months, state='readonly', width=12)
            month_combo.grid(row=0, column=1, padx=5, pady=5)
            month_combo.current(0)

            ttk.Label(select_frame, text="Year:").grid(row=0, column=2, padx=5, pady=5)
            from datetime import datetime
            year_entry = ttk.Entry(select_frame, width=6)
            year_entry.grid(row=0, column=3, padx=5, pady=5)
            year_entry.insert(0, str(datetime.now().year))

            def apply_filter():
                selected_month = month_combo.get()
                selected_year = year_entry.get().strip()

                if not selected_year:
                    messagebox.showwarning("Invalid Year", "Please enter a year.")
                    return

                target_month_year = f"{selected_month} {selected_year}"

                # Clear current selection
                log_listbox.selection_clear(0, tk.END)

                # Select matching logs
                matches = 0
                for idx, (med_name, month_year) in log_map.items():
                    if month_year == target_month_year:
                        log_listbox.selection_set(idx)
                        matches += 1

                month_dialog.destroy()

                if matches == 0:
                    messagebox.showinfo("No Matches", f"No logs found for {target_month_year}")
                else:
                    messagebox.showinfo("Selection Complete", f"Selected {matches} log(s) for {target_month_year}")

            btn_frame = ttk.Frame(month_dialog)
            btn_frame.pack(pady=20)
            ttk.Button(btn_frame, text="Select", command=apply_filter, width=12).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="Cancel", command=month_dialog.destroy, width=12).pack(side=tk.LEFT, padx=5)

        ttk.Button(selection_btn_frame, text="Select All", command=select_all_logs, width=15).pack(side=tk.LEFT, padx=3)
        ttk.Button(selection_btn_frame, text="Select by Month", command=select_by_month, width=15).pack(side=tk.LEFT, padx=3)
        ttk.Button(selection_btn_frame, text="Clear Selection", command=clear_selection, width=15).pack(side=tk.LEFT, padx=3)

        def on_patient_select(event):
            """Update log list when patient is selected"""
            log_listbox.delete(0, tk.END)
            log_map.clear()

            selection = patient_listbox.curselection()
            if not selection:
                return

            profile_id = patient_map[selection[0]]
            logs = self.log_manager.list_logs_for_profile(profile_id)

            for idx, (medicine_name, month_year) in enumerate(logs):
                display_text = f"{medicine_name} - {month_year}"
                log_listbox.insert(tk.END, display_text)
                log_map[idx] = (medicine_name, month_year)

        patient_listbox.bind('<<ListboxSelect>>', on_patient_select)

        # Export Options Section
        ttk.Label(dialog, text="3. Export Options:", font=self.HEADING_FONT).pack(pady=(15, 5))

        options_frame = ttk.Frame(dialog)
        options_frame.pack(padx=20, pady=5)

        export_method = tk.StringVar(value="extended_table")

        ttk.Radiobutton(
            options_frame,
            text="Extended Table (adds rows to same page)",
            variable=export_method,
            value="extended_table"
        ).pack(anchor=tk.W, pady=3)

        ttk.Radiobutton(
            options_frame,
            text="Continuation Pages (creates new pages)",
            variable=export_method,
            value="continuation_pages"
        ).pack(anchor=tk.W, pady=3)

        include_images = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            options_frame,
            text="Include medication card images (Word only)",
            variable=include_images
        ).pack(anchor=tk.W, pady=5)

        ttk.Label(options_frame, text="Export Formats:", font=self.HEADING_FONT).pack(anchor=tk.W, pady=(10, 5))

        format_checkboxes_frame = ttk.Frame(options_frame)
        format_checkboxes_frame.pack(anchor=tk.W, padx=20, pady=2)

        export_word = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            format_checkboxes_frame,
            text="Word (.docx)",
            variable=export_word
        ).pack(side=tk.LEFT, padx=5)

        export_pdf = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            format_checkboxes_frame,
            text="PDF (.pdf)",
            variable=export_pdf
        ).pack(side=tk.LEFT, padx=5)

        export_excel = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            format_checkboxes_frame,
            text="Excel (.xlsx)",
            variable=export_excel
        ).pack(side=tk.LEFT, padx=5)

        # Open folder after export checkbox
        open_folder_after = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            options_frame,
            text="Open folder after export",
            variable=open_folder_after
        ).pack(anchor=tk.W, pady=10)

        def proceed_with_export():
            """Execute the export"""
            # Validate selections
            patient_sel = patient_listbox.curselection()
            log_sel = log_listbox.curselection()

            if not patient_sel:
                messagebox.showwarning("No Patient Selected", "Please select a patient.")
                return

            if not log_sel:
                messagebox.showwarning("No Log Selected", "Please select at least one medication log.")
                return

            # Validate at least one format is selected
            if not export_word.get() and not export_pdf.get() and not export_excel.get():
                messagebox.showwarning("No Format Selected", "Please select at least one export format.")
                return

            profile_id = patient_map[patient_sel[0]]
            profile = self.profile_manager.get_profile(profile_id)

            dialog.destroy()

            # Ask user if they want to use default patient folder or custom
            use_default = messagebox.askyesno(
                "Export Location",
                "Use patient's default export folder?\n\n"
                "Yes: Save to patient folder (auto-organized)\n"
                "No: Choose custom location"
            )

            output_dir = None
            if not use_default:
                # Use default export folder from settings as initial directory
                initial_dir = self.settings_manager.get_default_export_folder() if self.settings_manager else str(Path.home() / "Documents")
                output_dir = filedialog.askdirectory(
                    title="Select Output Directory",
                    initialdir=initial_dir
                )
                if not output_dir:
                    return

            # Export each selected log
            all_results = []
            errors = []
            total_files = 0

            for log_idx in log_sel:
                medicine_name, month_year = log_map[log_idx]
                log_data = self.log_manager.get_log(profile_id, medicine_name, month_year)

                if not log_data:
                    errors.append(f"Failed to load: {medicine_name} - {month_year}")
                    continue

                try:
                    result = {}
                    log_files = []

                    # Export Word/PDF if requested
                    if export_word.get() or export_pdf.get():
                        word_pdf_result = self.export_manager.export_log(
                            profile,
                            log_data,
                            output_dir=output_dir,
                            create_pdf=export_pdf.get(),
                            export_method=export_method.get(),
                            include_images=include_images.get(),
                            data_dir=os.path.join(self.data_dir, "patients")
                        )
                        result.update(word_pdf_result)

                        if 'docx' in result:
                            log_files.append(os.path.basename(result['docx']))
                            total_files += 1
                        if 'pdf' in result:
                            log_files.append(os.path.basename(result['pdf']))
                            total_files += 1

                    # Export to Excel if requested
                    if export_excel.get():
                        # Determine Excel filename
                        if 'docx' in result:
                            excel_path = result['docx'].replace('.docx', '.xlsx')
                        else:
                            # Excel-only export, need to generate filename
                            if output_dir is None:
                                profile_id_str = profile.get('child_name', '').lower().replace(' ', '_')
                                output_dir = self.export_manager.get_patient_export_dir(profile_id_str, os.path.join(self.data_dir, "patients"))

                            medicine_name_clean = log_data.get('medicine_name', 'unknown').replace(' ', '_')
                            month_year_clean = log_data['month_year'].replace(' ', '_')
                            filename_base = f"{medicine_name_clean}_{month_year_clean}"
                            excel_path = os.path.join(output_dir, f"{filename_base}.xlsx")

                        try:
                            self._export_log_to_excel(log_data, excel_path)
                            result['xlsx'] = excel_path
                            log_files.append(os.path.basename(excel_path))
                            total_files += 1
                        except Exception as excel_error:
                            if export_word.get() or export_pdf.get():
                                errors.append(f"{medicine_name} - {month_year}: Excel export failed ({excel_error})")
                            else:
                                raise

                    all_results.append({
                        'log': f"{medicine_name} - {month_year}",
                        'files': log_files
                    })

                except Exception as e:
                    errors.append(f"{medicine_name} - {month_year}: {str(e)}")

            # Show results
            msg = f"Export Complete!\n\nExported {len(all_results)} log(s) to {total_files} file(s)\n\n"

            for item in all_results:
                msg += f"{item['log']}:\n"
                for file in item['files']:
                    msg += f"  • {file}\n"
                msg += "\n"

            if errors:
                msg += "\nErrors:\n"
                for error in errors:
                    msg += f"  ⚠ {error}\n"

            if errors:
                messagebox.showwarning("Export Complete with Errors", msg.strip())
            else:
                messagebox.showinfo("Export Complete", msg.strip())

            # Open folder if requested and files were exported
            if open_folder_after.get() and all_results:
                # Get the export directory from the first result
                # We need to trace back to find where the files were saved
                # The output_dir variable should be accessible here
                folder_to_open = output_dir

                # If output_dir is None, files went to patient folder
                if folder_to_open is None and all_results:
                    # Try to get directory from the first exported file
                    # We need to look at the actual result data
                    try:
                        # Get profile to find patient folder
                        profile_id_str = profile.get('child_name', '').lower().replace(' ', '_')
                        folder_to_open = self.export_manager.get_patient_export_dir(profile_id_str, os.path.join(self.data_dir, "patients"))
                    except:
                        folder_to_open = None

                if folder_to_open:
                    try:
                        system = platform.system()
                        if system == 'Windows':
                            os.startfile(folder_to_open)
                        elif system == 'Darwin':
                            subprocess.run(['open', folder_to_open], check=True)
                        else:
                            subprocess.run(['xdg-open', folder_to_open], check=True)
                    except Exception as folder_error:
                        # Silently fail - folder opening is a convenience feature
                        pass

        # Buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20, padx=10)

        ttk.Button(btn_frame, text="Export", command=proceed_with_export, width=15, padding=5).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy, width=15, padding=5).pack(side=tk.LEFT, padx=5)

    # Unsaved changes detection
    def check_unsaved_profile_changes(self):
        """Check for unsaved profile changes and prompt user"""
        if not self.profile_dirty:
            return True

        result = messagebox.askyesnocancel(
            "Unsaved Changes",
            "You have unsaved profile changes. Do you want to save them?",
            icon='warning'
        )

        if result is True:  # Yes - Save
            self.save_profile()
            return True
        elif result is False:  # No - Don't Save
            self.profile_dirty = False
            return True
        else:  # Cancel
            return False

    def check_unsaved_med_card_changes(self):
        """Check for unsaved medication card changes and prompt user"""
        if not self.med_card_dirty:
            return True

        result = messagebox.askyesnocancel(
            "Unsaved Changes",
            "You have unsaved medication card changes. Do you want to save them?",
            icon='warning'
        )

        if result is True:  # Yes - Save
            self.save_medication_card()
            return True
        elif result is False:  # No - Don't Save
            self.med_card_dirty = False
            return True
        else:  # Cancel
            return False

    def check_unsaved_log_changes(self):
        """Check for unsaved log changes and prompt user"""
        if not self.log_dirty:
            return True

        result = messagebox.askyesnocancel(
            "Unsaved Changes",
            "You have unsaved log changes. Do you want to save them?",
            icon='warning'
        )

        if result is True:  # Yes - Save
            self.save_medication_info()
            self.save_reason()
            return True
        elif result is False:  # No - Don't Save
            self.log_dirty = False
            return True
        else:  # Cancel
            return False

    def mark_profile_dirty(self, *args):
        """Mark profile as having unsaved changes"""
        self.profile_dirty = True

    def mark_med_card_dirty(self, *args):
        """Mark medication card as having unsaved changes"""
        self.med_card_dirty = True

    def mark_log_dirty(self, *args):
        """Mark log as having unsaved changes"""
        self.log_dirty = True

    # Other methods
    def show_about(self):
        """Show about dialog"""
        messagebox.showinfo("About",
                          "Medication Log Tracker\n\n"
                          "Manage patient profiles and medication administration logs\n\n"
                          "Data stored locally in JSON format")

    def open_data_folder(self):
        """Open the data folder in the system's file explorer"""
        data_dir = self.profile_manager.data_dir

        try:
            # Check if directory exists
            if not os.path.exists(data_dir):
                messagebox.showwarning("Folder Not Found",
                                      f"Data directory does not exist yet:\n{data_dir}\n\n"
                                      "It will be created when you save your first profile.")
                return

            # Open folder based on operating system
            system = platform.system()

            if system == 'Windows':
                # Windows: use explorer
                os.startfile(data_dir)
            elif system == 'Darwin':
                # macOS: use open command
                subprocess.run(['open', data_dir], check=True)
            else:
                # Linux: try xdg-open
                subprocess.run(['xdg-open', data_dir], check=True)

        except Exception as e:
            messagebox.showerror("Error Opening Folder",
                               f"Could not open data folder:\n{str(e)}\n\n"
                               f"Please navigate to:\n{data_dir}")

    def _create_settings_tab(self):
        """Create settings tab for application preferences"""
        # Main container with padding
        container = ttk.Frame(self.settings_tab)
        container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Title
        title_label = ttk.Label(container, text="Application Settings", font=self.TITLE_FONT)
        title_label.pack(pady=(0, 20))

        # Data Location Section (Read-only)
        data_frame = ttk.LabelFrame(container, text="Data Location", padding=15)
        data_frame.pack(fill=tk.X, pady=(0, 15))

        ttk.Label(data_frame, text="Current Data Location:", font=self.NORMAL_FONT).grid(row=0, column=0, sticky=tk.W, pady=5)

        data_location = self.settings_manager.get_data_location() if self.settings_manager else "Unknown"
        data_location_label = ttk.Label(data_frame, text=data_location, font=self.SMALL_FONT, foreground="gray")
        data_location_label.grid(row=1, column=0, sticky=tk.W, padx=20)

        ttk.Label(data_frame, text="All patient profiles, medication cards, and logs are stored here.",
                 font=self.SMALL_FONT, foreground="gray").grid(row=2, column=0, sticky=tk.W, pady=(5, 0))

        # Export Settings Section
        export_frame = ttk.LabelFrame(container, text="Export Settings", padding=15)
        export_frame.pack(fill=tk.X, pady=(0, 15))

        ttk.Label(export_frame, text="Default Export Folder:", font=self.NORMAL_FONT).grid(row=0, column=0, sticky=tk.W, pady=5)

        # Current default export folder
        current_export_folder = self.settings_manager.get_default_export_folder() if self.settings_manager else str(Path.home() / "Documents")

        export_folder_var = tk.StringVar(value=current_export_folder)
        export_folder_entry = ttk.Entry(export_frame, textvariable=export_folder_var, width=50, font=self.NORMAL_FONT)
        export_folder_entry.grid(row=1, column=0, sticky=tk.W, padx=20, pady=5)
        export_folder_entry.config(state='readonly')

        def browse_export_folder():
            """Open folder browser for export folder"""
            folder = filedialog.askdirectory(
                title="Select Default Export Folder",
                initialdir=export_folder_var.get()
            )
            if folder:
                export_folder_var.set(folder)
                save_export_settings()

        def save_export_settings():
            """Save the export folder setting"""
            if not self.settings_manager:
                messagebox.showerror("Error", "Settings manager not available")
                return

            folder = export_folder_var.get()
            if self.settings_manager.set_default_export_folder(folder):
                messagebox.showinfo("Settings Saved", "Default export folder updated successfully!")
            else:
                messagebox.showerror("Error", f"Failed to save settings.\n\nPlease ensure the folder exists:\n{folder}")

        btn_frame = ttk.Frame(export_frame)
        btn_frame.grid(row=2, column=0, sticky=tk.W, padx=20, pady=10)

        ttk.Button(btn_frame, text="Browse...", command=browse_export_folder, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Reset to Default",
                  command=lambda: export_folder_var.set(str(Path.home() / "Documents")),
                  width=15).pack(side=tk.LEFT, padx=5)

        ttk.Label(export_frame, text="This folder will be used as the default location when exporting medication logs.",
                 font=self.SMALL_FONT, foreground="gray").grid(row=3, column=0, sticky=tk.W, pady=(5, 0))

        # Info Section
        info_frame = ttk.Frame(container)
        info_frame.pack(fill=tk.X, pady=(20, 0))

        info_text = (
            "💡 Tips:\n"
            "• The data location is automatically set to your user folder and cannot be changed\n"
            "• The default export folder is used as the starting location when exporting logs\n"
            "• You can still choose a different location during each export if needed"
        )

        info_label = ttk.Label(info_frame, text=info_text, font=self.SMALL_FONT,
                              foreground="gray", justify=tk.LEFT)
        info_label.pack(anchor=tk.W)


def main():
    """Main entry point"""
    root = tk.Tk()
    app = MedicationTrackerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
